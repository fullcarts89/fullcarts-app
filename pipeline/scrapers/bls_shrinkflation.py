"""BLS R-CPI-SC (Research CPI excluding product size changes) scraper.

Downloads two quarterly-updated BLS Excel files and stores the parsed data
in the bls_shrinkflation table:

  r-cpi-sc-counts.xlsx  — Frequency of downsizing / upsizing by food category
  r-cpi-sc-data.xlsx    — Official CPI and R-CPI-SC indexes by food category

Data is updated quarterly (January, April, July, October).  The scraper
checks a minimum interval between downloads so we can schedule it monthly
without hammering BLS servers.

BLS Excel layout (typical):
  - First few rows may contain a title and blank rows (metadata)
  - A header row contains date labels like "Jan 2015", "Feb 2015", …
  - Subsequent rows are data rows: first cell = category name, rest = values
  - Some files split sections with a blank/label row between downsizing and
    upsizing blocks; the scraper detects this via keyword scanning.
"""
import io
import re
import tempfile
import time
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from pipeline.config import USER_AGENT
from pipeline.lib.cursor import load_cursor, save_cursor
from pipeline.lib.github_summary import write_summary
from pipeline.lib.http_client import RateLimitedSession
from pipeline.lib.logging_setup import get_logger
from pipeline.lib.supabase_client import get_client, reset_client
from pipeline.scrapers.base import BaseScraper

# ── Constants ──────────────────────────────────────────────────────────────

_COUNTS_URL = "https://www.bls.gov/cpi/research-series/r-cpi-sc-counts.xlsx"
_DATA_URL = "https://www.bls.gov/cpi/research-series/r-cpi-sc-data.xlsx"

# Skip re-download if already downloaded within this many days.
# BLS updates quarterly, so 25 days safely catches every update while
# preventing redundant downloads when the workflow runs monthly.
_MIN_DAYS_BETWEEN_DOWNLOADS = 25

# Be polite to BLS servers — one request every 5 seconds
_BLS_RPS = 0.2

# Retry config (mirrors BaseScraper)
_MAX_RETRIES = 3
_RETRY_BASE_DELAY = 5

# Month abbreviations for string-based date parsing
_MONTH_ABBR: Dict[str, int] = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Keywords that identify the "downsizing" vs "upsizing" section in the
# counts file (matched case-insensitively against sheet name or nearby cells)
_DOWNSIZE_KEYWORDS = ("downsize", "downsizing", "decrease", "reduction")
_UPSIZE_KEYWORDS = ("upsize", "upsizing", "increase")

# Keywords for official CPI vs R-CPI-SC series in the data file
_RCPI_SC_KEYWORDS = ("r-cpi-sc", "rcpi-sc", "research cpi", "without")


class BlsShrinkflationScraper(BaseScraper):
    """Downloads BLS R-CPI-SC counts and index data into bls_shrinkflation."""

    scraper_name = "bls_shrinkflation"
    source_type = "bls"  # not written to raw_items; kept for BaseScraper compat

    def __init__(self) -> None:
        super().__init__()
        self._session = RateLimitedSession(
            requests_per_second=_BLS_RPS,
            user_agent=USER_AGENT,
        )

    # ── BaseScraper interface ──────────────────────────────────────────────

    def fetch(
        self, cursor: Dict[str, Any], dry_run: bool = False
    ) -> List[Dict[str, Any]]:
        """Download and parse both BLS XLSX files.

        Returns a list of merged records keyed by (series, period).
        Each record may have counts data, index data, or both depending
        on whether the series appears in both files.

        Skips the download if already completed within _MIN_DAYS_BETWEEN_DOWNLOADS
        days (to avoid redundant hits on BLS servers during monthly runs).
        """
        last_downloaded = cursor.get("last_downloaded_at")
        if last_downloaded and not dry_run:
            last_dt = datetime.fromisoformat(last_downloaded)
            days_ago = (datetime.now(timezone.utc) - last_dt).days
            if days_ago < _MIN_DAYS_BETWEEN_DOWNLOADS:
                self.log.info(
                    "Skipping download — last run was %d days ago "
                    "(threshold: %d days)",
                    days_ago, _MIN_DAYS_BETWEEN_DOWNLOADS,
                )
                return []

        self.log.info("Downloading BLS XLSX files...")

        with tempfile.TemporaryDirectory() as tmpdir:
            counts_path = self._download_file(_COUNTS_URL, tmpdir, "counts.xlsx")
            data_path = self._download_file(_DATA_URL, tmpdir, "data.xlsx")

            if counts_path is None and data_path is None:
                self.log.error("Both BLS downloads failed — aborting")
                return []

            counts_records: Dict[Tuple[str, date], Dict[str, Any]] = {}
            data_records: Dict[Tuple[str, date], Dict[str, Any]] = {}

            if counts_path:
                counts_records = self._parse_counts_file(counts_path)
                self.log.info(
                    "Parsed %d series-period rows from counts file",
                    len(counts_records),
                )
            if data_path:
                data_records = self._parse_data_file(data_path)
                self.log.info(
                    "Parsed %d series-period rows from data file",
                    len(data_records),
                )

        # Merge on (series, period) — either file may have rows the other lacks
        merged: Dict[Tuple[str, date], Dict[str, Any]] = {}
        now_iso = datetime.now(timezone.utc).isoformat()

        for key, rec in counts_records.items():
            merged[key] = {
                "series": key[0],
                "period": key[1].isoformat(),
                "downsizing_count": rec.get("downsizing_count"),
                "upsizing_count": rec.get("upsizing_count"),
                "official_cpi": None,
                "rcpi_sc": None,
                "counts_url": _COUNTS_URL,
                "data_url": None,
                "downloaded_at": now_iso,
            }

        for key, rec in data_records.items():
            if key in merged:
                merged[key]["official_cpi"] = rec.get("official_cpi")
                merged[key]["rcpi_sc"] = rec.get("rcpi_sc")
                merged[key]["data_url"] = _DATA_URL
            else:
                merged[key] = {
                    "series": key[0],
                    "period": key[1].isoformat(),
                    "downsizing_count": None,
                    "upsizing_count": None,
                    "official_cpi": rec.get("official_cpi"),
                    "rcpi_sc": rec.get("rcpi_sc"),
                    "counts_url": None,
                    "data_url": _DATA_URL,
                    "downloaded_at": now_iso,
                }

        items = list(merged.values())
        self.log.info("Total merged records: %d", len(items))
        return items

    def source_id_for(self, item: Dict[str, Any]) -> str:
        return "bls_{}_{}".format(
            re.sub(r"\W+", "_", item["series"].lower()).strip("_"),
            item["period"],
        )

    def next_cursor(
        self, items: List[Dict[str, Any]], prev_cursor: Dict[str, Any]
    ) -> Dict[str, Any]:
        return {
            "last_downloaded_at": datetime.now(timezone.utc).isoformat(),
            "records_stored": len(items),
        }

    # ── Tall-format counts parser ──────────────────────────────────────────

    def _parse_tall_counts_sheet(
        self,
        rows: List[tuple],
        sheet_name: str,
    ) -> Dict[Tuple[str, date], Dict[str, Any]]:
        """Parse a tall-format BLS counts sheet.

        Layout (actual BLS r-cpi-sc-counts.xlsx as of 2025):
          Row 0: title row (ignored)
          Row 1: "Month" | "Downsize Count" | "Upsize Count"
          Row 2+: <datetime or date string> | <int> | <int>

        The series name is "All food" (aggregate total across all categories).
        """
        results: Dict[Tuple[str, date], Dict[str, Any]] = {}

        # Find header row: first row where col 0 is non-numeric text
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] is not None:
                cell = str(row[0]).strip().lower()
                if cell and not _is_numeric(row[0]) and _parse_period(row[0]) is None:
                    # Looks like a text label row — check for "count" in other cells
                    rest_labels = " ".join(
                        str(c).lower() for c in row[1:] if c is not None
                    )
                    if "count" in rest_labels or "downsize" in rest_labels:
                        header_idx = i
                        break

        if header_idx is None:
            self.log.warning(
                "Could not find header row in tall counts sheet '%s'",
                sheet_name,
            )
            return results

        header = [str(c).lower().strip() if c is not None else "" for c in rows[header_idx]]
        # Identify column indices
        downsize_col: Optional[int] = None
        upsize_col: Optional[int] = None
        for col_idx, label in enumerate(header):
            if "downsize" in label or "decrease" in label or "reduction" in label:
                downsize_col = col_idx
            elif "upsize" in label or "increase" in label:
                upsize_col = col_idx

        for row in rows[header_idx + 1:]:
            if not row or row[0] is None:
                continue
            period = _parse_period(row[0])
            if period is None:
                continue

            key = ("All food", period)
            vals: Dict[str, Any] = {}
            if downsize_col is not None and downsize_col < len(row):
                raw = row[downsize_col]
                if _is_numeric(raw):
                    vals["downsizing_count"] = int(round(float(raw)))
            if upsize_col is not None and upsize_col < len(row):
                raw = row[upsize_col]
                if _is_numeric(raw):
                    vals["upsizing_count"] = int(round(float(raw)))

            if vals:
                results[key] = vals

        return results

    # ── Custom store — writes to bls_shrinkflation, not raw_items ─────────

    def store(self, items: List[Dict[str, Any]]) -> int:
        """Upsert records into bls_shrinkflation.

        Uses ON CONFLICT (series, period) DO UPDATE so re-runs are idempotent
        and values are refreshed with each quarterly release.
        """
        client = get_client()
        total_upserted = 0
        batch_size = 50

        for i in range(0, len(items), batch_size):
            batch = items[i:i + batch_size]
            batch_num = i // batch_size

            resp = self._upsert_batch_with_retry(client, batch, batch_num)
            if resp and resp.data:
                total_upserted += len(resp.data)

            self.log.debug(
                "Batch %d-%d: upserted %d rows",
                i, i + len(batch),
                len(resp.data) if resp and resp.data else 0,
            )

        return total_upserted

    def _upsert_batch_with_retry(self, client, rows, batch_num):
        """Upsert a batch to bls_shrinkflation with exponential backoff."""
        for attempt in range(_MAX_RETRIES):
            try:
                return (
                    client.table("bls_shrinkflation")
                    .upsert(rows, on_conflict="series,period")
                    .execute()
                )
            except Exception as exc:
                exc_name = type(exc).__name__
                exc_str = str(exc)
                is_transient = (
                    "RemoteProtocolError" in exc_name
                    or "ConnectionTerminated" in exc_str
                    or "502" in exc_str
                    or "503" in exc_str
                    or "Bad gateway" in exc_str
                    or "Service Unavailable" in exc_str
                )
                if not is_transient:
                    raise
                delay = _RETRY_BASE_DELAY * (2 ** attempt)
                self.log.warning(
                    "Transient error at batch %d (attempt %d/%d): %s. "
                    "Retrying in %ds...",
                    batch_num, attempt + 1, _MAX_RETRIES, exc_name, delay,
                )
                time.sleep(delay)
                reset_client()
                client = get_client()

        # Final attempt
        return (
            client.table("bls_shrinkflation")
            .upsert(rows, on_conflict="series,period")
            .execute()
        )

    # ── File download ──────────────────────────────────────────────────────

    def _download_file(
        self, url: str, tmpdir: str, filename: str
    ) -> Optional[str]:
        """Download a BLS XLSX file to tmpdir.  Returns path or None."""
        import os
        dest = os.path.join(tmpdir, filename)
        self.log.info("Downloading %s → %s", url, filename)
        resp = self._session.get(url, raise_for_status=False)
        if resp is None or not resp.ok:
            self.log.error(
                "Failed to download %s: %s",
                url,
                resp.status_code if resp else "no response",
            )
            return None
        with open(dest, "wb") as f:
            f.write(resp.content)
        self.log.info(
            "Downloaded %s (%d bytes)", filename, len(resp.content)
        )
        return dest

    # ── COUNTS file parser ─────────────────────────────────────────────────

    def _parse_counts_file(
        self, path: str
    ) -> Dict[Tuple[str, date], Dict[str, Any]]:
        """Parse r-cpi-sc-counts.xlsx.

        The counts file tracks how many items in each food category were
        downsized or upsized in each month.  It may use:
          - Separate sheets for downsizing vs upsizing
          - A single sheet with two stacked blocks separated by a label row
          - Column pairs (downsizing | upsizing) for each period

        This parser scans the workbook structure to determine which layout
        is in use, then dispatches to the appropriate sub-parser.
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        results: Dict[Tuple[str, date], Dict[str, Any]] = {}

        self.log.info(
            "Counts file sheets: %s", wb.sheetnames
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lower = sheet_name.lower()

            # Determine if this sheet represents downsizing, upsizing, or both
            if any(k in sheet_lower for k in _DOWNSIZE_KEYWORDS):
                metric = "downsizing"
            elif any(k in sheet_lower for k in _UPSIZE_KEYWORDS):
                metric = "upsizing"
            else:
                metric = "both"  # will detect via in-sheet labels

            self.log.info(
                "Parsing counts sheet '%s' (metric=%s)", sheet_name, metric
            )

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Detect format: if the header row has a date in col 0 and
            # "count"-style column names, it's tall format (each row = one month).
            # Otherwise fall back to wide format (each column = one month).
            if _is_tall_counts_sheet(rows):
                self.log.info(
                    "  → detected tall format for '%s'", sheet_name
                )
                sheet_results = self._parse_tall_counts_sheet(rows, sheet_name)
            else:
                sheet_results = self._parse_wide_sheet(
                    rows, sheet_name, metric=metric
                )
            self.log.info(
                "  → found %d (series, period) rows in '%s'",
                len(sheet_results), sheet_name,
            )

            for key, vals in sheet_results.items():
                if key not in results:
                    results[key] = {}
                results[key].update(vals)

        wb.close()
        return results

    # ── DATA file parser ───────────────────────────────────────────────────

    def _parse_data_file(
        self, path: str
    ) -> Dict[Tuple[str, date], Dict[str, Any]]:
        """Parse r-cpi-sc-data.xlsx.

        The data file contains CPI index values (base = December 2014).
        Each food category typically appears twice: once for the standard
        CPI and once for the R-CPI-SC (research CPI excluding size changes).

        Rows labelled with "R-CPI-SC" or "research" keywords are treated as
        the adjusted series; all others are treated as official CPI.
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        results: Dict[Tuple[str, date], Dict[str, Any]] = {}

        self.log.info("Data file sheets: %s", wb.sheetnames)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            self.log.info("Parsing data sheet '%s'", sheet_name)
            sheet_results = self._parse_wide_sheet(
                rows, sheet_name, metric="index"
            )
            self.log.info(
                "  → found %d (series, period) rows in '%s'",
                len(sheet_results), sheet_name,
            )

            for key, vals in sheet_results.items():
                if key not in results:
                    results[key] = {}
                results[key].update(vals)

        wb.close()
        return results

    # ── Core wide-format sheet parser ─────────────────────────────────────

    def _parse_wide_sheet(
        self,
        rows: List[tuple],
        sheet_name: str,
        metric: str,
    ) -> Dict[Tuple[str, date], Dict[str, Any]]:
        """Parse a wide-format BLS sheet (categories as rows, dates as columns).

        Args:
            rows:       All rows from ws.iter_rows(values_only=True)
            sheet_name: Used for logging and heuristics
            metric:     "downsizing" | "upsizing" | "both" | "index"
                        Controls which result key(s) are populated.

        Returns:
            Dict keyed by (series_name, period_date) → {metric_key: value}
        """
        results: Dict[Tuple[str, date], Dict[str, Any]] = {}

        # Step 1: find the header row — the first row that has ≥3 parseable
        # date values in its cells (skips title/blank rows at the top)
        header_row_idx = _find_header_row(rows)
        if header_row_idx is None:
            self.log.warning(
                "Could not find date header row in sheet '%s' — skipping",
                sheet_name,
            )
            return results

        header = rows[header_row_idx]
        self.log.debug(
            "Header row %d in '%s': %s…",
            header_row_idx, sheet_name, str(header[:6]),
        )

        # Step 2: build period → column-index mapping from the header row.
        # The first column is the series name; remaining columns are periods.
        period_cols: List[Tuple[int, date]] = []
        for col_idx, cell_val in enumerate(header):
            if col_idx == 0:
                continue  # series name column
            parsed = _parse_period(cell_val)
            if parsed is not None:
                period_cols.append((col_idx, parsed))

        if not period_cols:
            self.log.warning(
                "No date columns found in header row %d of '%s'",
                header_row_idx, sheet_name,
            )
            return results

        self.log.debug(
            "'%s': found %d period columns (%s … %s)",
            sheet_name,
            len(period_cols),
            period_cols[0][1].isoformat(),
            period_cols[-1][1].isoformat(),
        )

        # Step 3: determine if we need to detect in-sheet section breaks
        # (e.g. counts file with a "Downsizing" block then "Upsizing" block)
        current_metric = metric  # may change when we hit a section-label row

        # Step 4: parse data rows
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            if not row or row[0] is None:
                continue  # blank row

            series_raw = str(row[0]).strip()
            if not series_raw:
                continue

            # Detect in-sheet section-label rows (no numeric data, just a label)
            lower_series = series_raw.lower()
            if any(k in lower_series for k in _DOWNSIZE_KEYWORDS):
                current_metric = "downsizing"
                self.log.debug(
                    "Row %d: switched to downsizing section ('%s')",
                    row_idx, series_raw,
                )
                continue
            if any(k in lower_series for k in _UPSIZE_KEYWORDS):
                current_metric = "upsizing"
                self.log.debug(
                    "Row %d: switched to upsizing section ('%s')",
                    row_idx, series_raw,
                )
                continue

            # Check if this row has any numeric values in period columns
            has_data = any(
                _is_numeric(row[col_idx])
                for col_idx, _ in period_cols
                if col_idx < len(row)
            )
            if not has_data:
                continue  # subtotal / footnote / blank data row

            # Detect R-CPI-SC vs official CPI for index sheets
            if current_metric == "index":
                is_rcpi_sc = any(k in lower_series for k in _RCPI_SC_KEYWORDS)
                # Strip the R-CPI-SC qualifier from the series name for matching
                clean_series = _clean_series_name(series_raw, is_rcpi_sc)
            else:
                is_rcpi_sc = False
                clean_series = series_raw

            # Extract values for each period
            for col_idx, period in period_cols:
                if col_idx >= len(row):
                    continue
                raw_val = row[col_idx]
                if not _is_numeric(raw_val):
                    continue
                value = float(raw_val)
                key = (clean_series, period)
                if key not in results:
                    results[key] = {}

                if current_metric == "downsizing":
                    results[key]["downsizing_count"] = int(round(value))
                elif current_metric == "upsizing":
                    results[key]["upsizing_count"] = int(round(value))
                elif current_metric == "both":
                    # Counts file with no section break detected — store under
                    # a generic count key; caller can rename if needed
                    results[key].setdefault("count", int(round(value)))
                elif current_metric == "index":
                    if is_rcpi_sc:
                        results[key]["rcpi_sc"] = round(value, 3)
                    else:
                        results[key]["official_cpi"] = round(value, 3)

        return results


# ── Module-level helpers ───────────────────────────────────────────────────

def _is_tall_counts_sheet(rows: List[tuple]) -> bool:
    """Return True if the sheet is in tall format (one month per row).

    Tall format: has a text header row with "count" labels, and data rows
    where the first cell is a date.  Contrast with wide format where dates
    appear as column headers.
    """
    for row in rows[:5]:
        if not row or row[0] is None:
            continue
        rest = " ".join(str(c).lower() for c in row[1:] if c is not None)
        if "count" in rest or "downsize" in rest:
            return True
    return False


def _parse_period(cell_value: Any) -> Optional[date]:
    """Parse a BLS period cell into a date (first of month).

    Handles:
      - Python datetime / date objects (openpyxl date-formatted cells)
      - "Jan 2015"  / "Jan-2015"  / "Jan-15"
      - "2015 Jan"  / "2015-01"
      - "2015M01"   (BLS internal format)
    """
    if cell_value is None:
        return None

    if isinstance(cell_value, datetime):
        return cell_value.date().replace(day=1)
    if isinstance(cell_value, date):
        return cell_value.replace(day=1)

    s = str(cell_value).strip()
    if not s:
        return None

    # "YYYY-MM"
    m = re.match(r"^(\d{4})-(\d{2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), 1)
        except ValueError:
            pass

    # "YYYYMNN" BLS internal e.g. "2015M01"
    m = re.match(r"^(\d{4})M(\d{2})$", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), 1)
        except ValueError:
            pass

    # "Mon YYYY" or "Mon-YYYY" or "Mon-YY"
    m = re.match(r"^([A-Za-z]{3})[-\s](\d{2,4})$", s)
    if m:
        month = _MONTH_ABBR.get(m.group(1).lower())
        if month:
            year = int(m.group(2))
            if year < 100:
                year += 2000
            try:
                return date(year, month, 1)
            except ValueError:
                pass

    # "YYYY Mon"
    m = re.match(r"^(\d{4})\s+([A-Za-z]{3})$", s)
    if m:
        month = _MONTH_ABBR.get(m.group(2).lower())
        if month:
            try:
                return date(int(m.group(1)), month, 1)
            except ValueError:
                pass

    return None


def _is_numeric(val: Any) -> bool:
    """Return True if val is a non-None number."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val))
        return True
    except (ValueError, TypeError):
        return False


def _find_header_row(rows: List[tuple]) -> Optional[int]:
    """Return the index of the first row containing ≥3 parseable date cells."""
    for idx, row in enumerate(rows):
        date_count = sum(
            1 for cell in row
            if cell is not None and _parse_period(cell) is not None
        )
        if date_count >= 3:
            return idx
    return None


def _clean_series_name(series: str, is_rcpi_sc: bool) -> str:
    """Strip R-CPI-SC qualifiers from a series name for cross-file matching.

    e.g. "All food - R-CPI-SC" → "All food"
    """
    if not is_rcpi_sc:
        return series.strip()
    # Remove common suffixes/prefixes used to label the adjusted series
    for pat in (
        r"\s*[-–—]\s*R-CPI-SC.*$",
        r"\s*[-–—]\s*Research CPI.*$",
        r"\s*\(R-CPI-SC\).*$",
        r"\s*\(without.*\).*$",
    ):
        cleaned = re.sub(pat, "", series, flags=re.IGNORECASE).strip()
        if cleaned and cleaned != series:
            return cleaned
    return series.strip()
